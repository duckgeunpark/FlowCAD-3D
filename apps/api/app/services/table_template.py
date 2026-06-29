"""Canonical input-table columns + Excel/CSV transport normalization."""
from __future__ import annotations

import io

from ..domain.enums import DesignMode
from ..parsing.base import Row

# This is the backend-owned transport schema. The frontend table mirrors this
# list until the project grows a generated shared schema.
_ASSEMBLY_COLUMNS = [
    "seq", "system_type", "part_type", "spec",
    "size_a", "size_b", "length", "angle",
    "bend_to", "offset_direction", "rotation",
    "W", "H", "D", "L", "R",
    "toW", "toH", "toD",
    "branchW", "branchH", "branchD",
    "offset", "X", "NL", "gores",
    "connect_to_seq", "connect_port", "note",
]
PIPE_COLUMNS = _ASSEMBLY_COLUMNS
DUCT_COLUMNS = _ASSEMBLY_COLUMNS

# --- v2 duct schema (``duct_3d_sheet_v2.xlsx`` / Duct3D_Input) --------------- #
V2_SHEET_NAME = "Duct3D_Input"
V2_DUCT_COLUMNS = [
    "row_type", "seq", "line_id", "system_id", "service", "level_name", "zone",
    "element_id", "parent_element_id", "from_element_id", "to_element_id",
    "branch_to_element_id", "element_type", "family_code", "standard_code",
    "shape_code", "material_code", "gauge_t", "insulation_t", "lining_t",
    "spec_code", "origin_mode", "origin_x", "origin_y", "origin_z",
    "end_x", "end_y", "end_z", "dir_x", "dir_y", "dir_z", "up_x", "up_y", "up_z",
    "orientation_code", "mirror_code", "rotation_deg", "slope_ratio",
    "centerline_length", "path_length", "width", "height", "diameter",
    "major_axis", "minor_axis", "inlet_width", "inlet_height", "inlet_diameter",
    "outlet_width", "outlet_height", "outlet_diameter", "branch_width",
    "branch_height", "branch_diameter", "fitting_type", "angle_deg",
    "branch_angle_deg", "part_subtype", "branch_plane_code", "branch_b_side",
    "branch_c_side", "radius_type", "centerline_radius", "throat_radius",
    "heel_radius", "offset_x", "offset_y", "offset_z", "eccentric_side",
    "transition_length", "taper_method", "reducer_method", "tee_run_length",
    "tee_branch_length", "tap_location_ratio", "vane_count", "port_a_role",
    "port_a_dir", "port_b_role", "port_b_dir", "port_c_role", "port_c_dir",
    "port_d_role", "port_d_dir", "main_flow_from", "main_flow_to", "joint_a_type",
    "joint_b_type", "joint_c_type", "joint_d_type", "flange_type",
    "connector_code", "bom_item_no", "part_name_ko", "part_name_en", "qty",
    "unit", "calc_rule", "weight_kg", "surface_area_m2", "smacna_code", "note",
    "ai_source", "ai_confidence", "review_status", "error_code", "error_message",
]

# One example DATA row (a 500x300 rectangular straight) for the duct template.
_V2_DUCT_EXAMPLE = {
    "row_type": "DATA", "seq": 10, "line_id": "L-001", "system_id": "SA-01",
    "service": "SA", "level_name": "L1", "zone": "Z1", "element_id": "E0001",
    "to_element_id": "E0002", "element_type": "STRAIGHT",
    "family_code": "STRAIGHT_RECT", "shape_code": "RECT", "material_code": "GI",
    "gauge_t": 0.8, "spec_code": "STD-A", "origin_mode": "CENTER",
    "origin_x": 0, "origin_y": 0, "origin_z": 3000, "end_x": 2000, "end_y": 0,
    "end_z": 3000, "dir_x": 1, "dir_y": 0, "dir_z": 0, "up_z": 1,
    "orientation_code": "XP_XP", "centerline_length": 2000, "width": 500,
    "height": 300, "fitting_type": "NONE", "port_a_role": "INLET",
    "port_a_dir": "IN", "port_b_role": "OUTLET", "port_b_dir": "OUT",
    "bom_item_no": "001", "part_name_ko": "직관", "part_name_en": "Straight Duct",
    "qty": 1, "unit": "EA", "review_status": "APPROVED",
}

# Field requirement flags for the spec row (mirrors the sheet's REQ/OPT row).
_V2_REQUIRED = {
    "row_type", "seq", "element_id", "element_type", "family_code", "shape_code",
    "origin_x", "origin_y", "origin_z", "dir_x", "dir_y", "dir_z",
    "up_x", "up_y", "up_z", "orientation_code", "fitting_type",
    "port_a_role", "port_b_role", "part_name_en", "review_status",
}


def is_v2_header(header: list[str]) -> bool:
    """True when a parsed sheet's header is the v2 duct schema."""
    keys = {h.strip().lower() for h in header}
    return "element_id" in keys and ("family_code" in keys or "element_type" in keys)


def _is_v2_rows(rows: list[Row]) -> bool:
    return bool(rows) and is_v2_header(list(rows[0].keys()))

_PIPE_EXAMPLE = [
    1, "pipe", "straight", "SCH40", 100, "", 3000, "",
    "", "", "", "", "", "", "", "", "", "", "", "", "", "",
    "", "", "", "", "", "start", "start pipe",
]
_DUCT_EXAMPLE = [
    1, "duct", "rect_straight", "GI", "", "", "", "",
    "", "", "", 500, 300, "", 1220, "", "", "", "", "", "",
    "", "", "", "", "", "", "start", "start duct",
]

_HEADER_ALIASES = {
    "sequence": "seq",
    "no": "seq",
    "번호": "seq",
    "순번": "seq",
    "계통": "system_type",
    "시스템": "system_type",
    "종류": "part_type",
    "부품종류": "part_type",
    "피팅": "part_type",
    "표준피팅": "part_type",
    "규격": "spec",
    "규격코드": "spec",
    "치수a": "size_a",
    "치수b": "size_b",
    "가로": "W",
    "폭": "W",
    "세로": "H",
    "높이": "H",
    "직경": "D",
    "지름": "D",
    "길이": "L",
    "반경": "R",
    "출구가로": "toW",
    "출구세로": "toH",
    "출구직경": "toD",
    "분기가로": "branchW",
    "분기세로": "branchH",
    "분기직경": "branchD",
    "각도": "angle",
    "엘보방향": "bend_to",
    "방향": "bend_to",
    "오프셋방향": "offset_direction",
    "회전": "rotation",
    "연결대상": "connect_to_seq",
    "연결seq": "connect_to_seq",
    "연결포트": "connect_port",
    "비고": "note",
}


def columns_for(mode: DesignMode) -> list[str]:
    return PIPE_COLUMNS if mode is DesignMode.PIPE else DUCT_COLUMNS


def normalize_table_rows(rows: list[Row]) -> list[Row]:
    """Normalize parsed transport rows into canonical design-table rows.

    v2 duct rows (``element_id`` + ``family_code``/``element_type`` schema) are
    passed through verbatim — they are already the canonical engine format, so the
    legacy assembly header-aliasing must not rename their columns.
    """
    if _is_v2_rows(rows):
        return [dict(r) for r in rows]
    normalized: list[Row] = []
    for raw in rows:
        row: Row = {}
        for key, value in raw.items():
            canonical = _canonical_header(key)
            if canonical:
                row[canonical] = value
        _fill_compat_dimensions(row)
        normalized.append(row)
    return normalized


def build_template_xlsx(mode: DesignMode) -> bytes:
    """Empty workbook with a bold header row, a frozen header, and one example."""
    if mode is DesignMode.DUCT:
        return _build_v2_duct_template()

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    columns = columns_for(mode)
    wb = Workbook()
    ws = wb.active
    ws.title = "FlowCAD"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="44607A")
    for col_idx, name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        ws.column_dimensions[cell.column_letter].width = max(10, len(name) + 2)

    for col_idx, value in enumerate(_PIPE_EXAMPLE, start=1):
        ws.cell(row=2, column=col_idx, value=value)

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_v2_duct_template() -> bytes:
    """v2 duct template: header row, a REQ/OPT spec row, then one example DATA row."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = V2_SHEET_NAME

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="44607A")
    for col_idx, name in enumerate(V2_DUCT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        ws.column_dimensions[cell.column_letter].width = max(10, len(name) + 2)
        # Spec row (row 2): REQ/OPT per field.
        ws.cell(row=2, column=col_idx,
                value="REQ" if name in _V2_REQUIRED else "OPT")
        ws.cell(row=3, column=col_idx, value=_V2_DUCT_EXAMPLE.get(name, ""))

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def load_table(filename: str, data: bytes) -> list[Row]:
    """Parse an uploaded .xlsx/.csv and normalize it for table/generation use."""
    if filename.lower().endswith((".xlsx", ".xlsm")):
        return normalize_table_rows(_load_xlsx(data))
    from .csv_loader import load_csv
    return normalize_table_rows(load_csv(data))


def _load_xlsx(data: bytes) -> list[Row]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    # Prefer the v2 duct input sheet when the workbook carries it (the v2 export
    # also ships CodeLists / ReadMe sheets we must ignore).
    ws = wb[V2_SHEET_NAME] if V2_SHEET_NAME in wb.sheetnames else wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = [str(h).strip() if h is not None else "" for h in next(rows_iter)]
    except StopIteration:
        return []

    is_v2 = is_v2_header(header)
    rows: list[Row] = []
    for values in rows_iter:
        if values is None or all(v is None or v == "" for v in values):
            continue
        row: Row = {}
        for key, value in zip(header, values):
            if key:
                row[key] = "" if value is None else value
        # v2 sheets carry a REQ/OPT spec row right under the header — skip it.
        if is_v2 and _is_v2_spec_row(row):
            continue
        rows.append(row)
    return rows


def _is_v2_spec_row(row: Row) -> bool:
    """The v2 spec row marks every field as REQ/OPT instead of carrying data."""
    rt = str(row.get("row_type", "")).strip().upper()
    if rt in ("REQ", "OPT"):
        return True
    values = {str(v).strip().upper() for v in row.values() if str(v).strip()}
    return bool(values) and values.issubset({"REQ", "OPT"})


def _canonical_header(header: object) -> str:
    raw = str(header or "").strip()
    if not raw:
        return ""
    if raw in _ASSEMBLY_COLUMNS:
        return raw
    compact = raw.lower().replace(" ", "").replace("_", "").replace("-", "")
    for col in _ASSEMBLY_COLUMNS:
        if compact == col.lower().replace("_", ""):
            return col
    return _HEADER_ALIASES.get(compact, raw)


def _fill_compat_dimensions(row: Row) -> None:
    """Populate legacy size columns for existing table and assembly logic."""
    part_type = str(row.get("part_type", "")).strip().lower()
    if row.get("length") in (None, "") and row.get("L") not in (None, ""):
        row["length"] = row["L"]
    if row.get("size_a") not in (None, ""):
        return
    if part_type.startswith("transition"):
        if row.get("toD") not in (None, ""):
            row["size_a"] = row["toD"]
        elif row.get("toW") not in (None, ""):
            row["size_a"] = row["toW"]
        if row.get("size_b") in (None, "") and row.get("toH") not in (None, ""):
            row["size_b"] = row["toH"]
        return
    if row.get("W") not in (None, ""):
        row["size_a"] = row["W"]
        if row.get("size_b") in (None, "") and row.get("H") not in (None, ""):
            row["size_b"] = row["H"]
    elif row.get("D") not in (None, ""):
        row["size_a"] = row["D"]
